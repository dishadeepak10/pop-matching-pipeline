import json
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE = Path(".")
OUTPUT = BASE / "data" / "output"

cases = [
    "00084283","00084285","00084308","00084323","00084360",
    "00084362","00084373","00084375","00084379","00084384",
    "00084401","00084501","00084572","00084596","00084670",
    "00084696","00084725","00084741","00084742","00084772",
    "00084822","00084826","00084851","00084879","00084922"
]

date_patterns = re.compile(
    r"(date|time|submitted|available|receipt|payment|wire|effective|"
    r"transaction|value|delivery|printed)",
    re.IGNORECASE
)

rows = []

for case in cases:

    json_path = OUTPUT / f"{case}_POP_Document" / "extracted.json"
    ocr_path = OUTPUT / f"{case}_POP_Document" / "ocr.txt"

    extracted_date = ""
    confidence = ""
    field_name = ""
    ocr_evidence = ""
    status = ""

    # ---------------------------------------------------------
    # Read extracted JSON
    # ---------------------------------------------------------
    if json_path.exists():
        data = json.loads(json_path.read_text(encoding="utf-8"))

        fields = data.get("fields", [])

        date_fields = [
            f for f in fields
            if date_patterns.search(str(f.get("field_name", "")))
            and "date_of_birth" not in str(f.get("field_name", "")).lower()
        ]

        # Prefer the same field families used by normalize_record()
        preferred = [
            "transaction_date",
            "transfer_date",
            "payment_date",
            "receipt_date",
            "date",
            "date_of_application",
            "wire_date",
            "effective_date",
            "value_date",
            "date_and_time",
            "transfer_date_time",
            "transaction_date_time",
            "transaction_time",
            "request_submitted_date_time",
            "payment_in_progress_date_time",
            "payment_submitted_date_time",
            "available_date_time_aest",
            "available_date_time_gst",
            "receipt_1_receipt_date",
            "receipt_2_receipt_date",
            "declaration_signature_date",
        ]

        for preferred_name in preferred:
            matches = [
                f for f in fields
                if str(f.get("field_name", "")).lower() == preferred_name.lower()
            ]

            if matches:
                value = matches[0].get("value")
                if value not in (None, "", "N/A", "Not Applicable"):
                    field_name = preferred_name
                    extracted_date = str(value)
                    confidence = matches[0].get("confidence", "")
                    break

    # ---------------------------------------------------------
    # Find matching evidence in OCR
    # ---------------------------------------------------------
    if ocr_path.exists():
        ocr_lines = ocr_path.read_text(
            encoding="utf-8",
            errors="replace"
        ).splitlines()

        evidence_lines = []

        for line in ocr_lines:
            line = line.strip()

            if not line:
                continue

            if date_patterns.search(line):
                evidence_lines.append(line)

        # Keep the evidence concise
        ocr_evidence = " | ".join(evidence_lines[:5])

    # ---------------------------------------------------------
    # Validation status
    # ---------------------------------------------------------
    if extracted_date:
        status = "EXTRACTED / OCR EVIDENCE AVAILABLE"
    elif ocr_evidence:
        status = "DATE NOT NORMALIZED - OCR DATE/TIME PRESENT"
    else:
        status = "NOT AVAILABLE IN POP/OCR"

    rows.append({
        "case_number": case,
        "transaction_date": extracted_date,
        "source_field": field_name,
        "confidence": confidence,
        "date_source": "POP" if extracted_date else "Not available in POP/OCR",
        "ocr_evidence": ocr_evidence,
        "validation_status": status,
    })

# -------------------------------------------------------------
# Create Excel
# -------------------------------------------------------------
df = pd.DataFrame(rows)

output_file = OUTPUT / "POP_date_validation.xlsx"

df.to_excel(output_file, index=False, sheet_name="Date Validation")

# -------------------------------------------------------------
# Formatting
# -------------------------------------------------------------
wb = load_workbook(output_file)
ws = wb["Date Validation"]

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter

    for cell in column:
        if cell.value is not None:
            max_length = max(max_length, len(str(cell.value)))

    ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

# Highlight validation status
for row in range(2, ws.max_row + 1):
    status_cell = ws.cell(row=row, column=7)

    if "EXTRACTED" in str(status_cell.value):
        status_cell.fill = PatternFill(
            fill_type="solid",
            fgColor="C6EFCE"
        )
    elif "NOT AVAILABLE" in str(status_cell.value):
        status_cell.fill = PatternFill(
            fill_type="solid",
            fgColor="FFF2CC"
        )

wb.save(output_file)

print("=" * 70)
print("DATE VALIDATION EXCEL CREATED")
print("=" * 70)
print(f"Total cases : {len(df)}")
print(f"Extracted   : {(df['transaction_date'] != '').sum()}")
print(f"Unavailable : {(df['transaction_date'] == '').sum()}")
print(f"Output      : {output_file}")
