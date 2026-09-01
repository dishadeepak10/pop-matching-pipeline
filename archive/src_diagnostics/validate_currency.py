import json
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


BASE = Path(".")
OUTPUT = BASE / "data" / "output"

cases = [
    "00084283", "00084285", "00084308", "00084323", "00084360",
    "00084362", "00084373", "00084375", "00084379", "00084384",
    "00084401", "00084501", "00084572", "00084596", "00084670",
    "00084696", "00084725", "00084741", "00084742", "00084772",
    "00084822", "00084826", "00084851", "00084879", "00084922"
]


VALID_CURRENCIES = {
    "AED", "AUD", "CAD", "CHF", "CNY",
    "EUR", "GBP", "HKD", "INR", "JPY",
    "NZD", "SAR", "SGD", "USD", "USDT"
}


currency_pattern = re.compile(
    r"(?<![A-Z])("
    r"AED|AUD|CAD|CHF|CNY|EUR|GBP|HKD|INR|JPY|NZD|"
    r"SAR|SGD|USD|USDT"
    r")(?![A-Z])",
    re.IGNORECASE
)


symbol_currency_pattern = re.compile(r"[$€£]")


# Fields that should NOT be treated as the primary currency source.
EXCLUDED_FIELDS = {
    "terms_and_conditions",
    "notes",
    "remarks",
    "description",
    "narration",
    "comments",
}


# Field names that are strong indicators of actual
# currency / monetary information.
PRIMARY_CURRENCY_PATTERN = re.compile(
    r"(currency|amount|fee|charge|total|"
    r"transfer_amount|payment_amount|transaction_amount|"
    r"debit_amount|credit_amount|"
    r"wire_amount|original_amount|"
    r"receipt.*amount)",
    re.IGNORECASE
)


# Exchange-rate fields contain currency information,
# but they are supporting evidence rather than primary currency.
EXCHANGE_RATE_PATTERN = re.compile(
    r"(exchange.?rate|conversion.?rate)",
    re.IGNORECASE
)


rows = []


for case in cases:

    json_path = OUTPUT / f"{case}_POP_Document" / "extracted.json"
    ocr_path = OUTPUT / f"{case}_POP_Document" / "ocr.txt"

    primary_values = []
    primary_fields = []
    primary_confidences = []

    evidence_values = []
    evidence_fields = []

    primary_currencies = set()
    evidence_currencies = set()

    ocr_evidence_lines = []

    # ---------------------------------------------------------
    # Read extracted JSON
    # ---------------------------------------------------------
    if json_path.exists():

        try:
            data = json.loads(
                json_path.read_text(encoding="utf-8")
            )

            fields = data.get("fields", [])

            if isinstance(fields, list):

                for field in fields:

                    if not isinstance(field, dict):
                        continue

                    field_name = str(
                        field.get("field_name", "")
                    ).strip()

                    field_name_lower = field_name.lower()

                    value = field.get("value")

                    if value in (
                        None,
                        "",
                        "N/A",
                        "Not Applicable"
                    ):
                        continue

                    value_text = str(value).strip()

                    currencies = {
                        match.upper()
                        for match in currency_pattern.findall(
                            value_text
                        )
                    }

                    has_symbol = bool(
                        symbol_currency_pattern.search(value_text)
                    )

                    if not currencies and not has_symbol:
                        continue

                    # -------------------------------------------------
                    # Ignore incidental free-text fields as primary
                    # currency sources.
                    # -------------------------------------------------
                    if field_name_lower in EXCLUDED_FIELDS:
                        evidence_values.append(value_text)
                        evidence_fields.append(field_name)

                        evidence_currencies.update(currencies)
                        continue

                    # -------------------------------------------------
                    # Exchange/conversion rates are supporting evidence.
                    # -------------------------------------------------
                    if EXCHANGE_RATE_PATTERN.search(field_name):

                        evidence_values.append(value_text)
                        evidence_fields.append(field_name)

                        evidence_currencies.update(currencies)
                        continue

                    # -------------------------------------------------
                    # Primary currency/amount information.
                    # -------------------------------------------------
                    if PRIMARY_CURRENCY_PATTERN.search(field_name):

                        primary_values.append(value_text)
                        primary_fields.append(field_name)

                        confidence = field.get(
                            "confidence",
                            ""
                        )

                        primary_confidences.append(confidence)

                        primary_currencies.update(currencies)

                    else:

                        # Currency found in another field:
                        # keep it as supporting evidence.
                        evidence_values.append(value_text)
                        evidence_fields.append(field_name)

                        evidence_currencies.update(currencies)

        except Exception as e:

            rows.append({
                "case_number": case,
                "primary_currency_values": "",
                "primary_source_fields": "",
                "detected_currencies": "",
                "confidence": "",
                "currency_evidence_fields": "",
                "currency_evidence": "",
                "ocr_evidence": "",
                "validation_status": f"ERROR READING JSON: {e}",
            })

            continue

    # ---------------------------------------------------------
    # OCR currency evidence
    # ---------------------------------------------------------
    if ocr_path.exists():

        try:
            ocr_lines = ocr_path.read_text(
                encoding="utf-8",
                errors="replace"
            ).splitlines()

            for line in ocr_lines:

                line = line.strip()

                if not line:
                    continue

                if (
                    currency_pattern.search(line)
                    or symbol_currency_pattern.search(line)
                ):
                    ocr_evidence_lines.append(line)

            ocr_evidence_lines = ocr_evidence_lines[:5]

        except Exception:
            pass

    ocr_evidence = " | ".join(
        ocr_evidence_lines
    )

    # ---------------------------------------------------------
    # Validation
    # ---------------------------------------------------------
    if primary_currencies:

        invalid_currencies = sorted(
            currency
            for currency in primary_currencies
            if currency not in VALID_CURRENCIES
        )

        if invalid_currencies:
            validation_status = "INVALID CURRENCY CODE"
        else:
            validation_status = "VALID CURRENCY"

    elif primary_values:

        validation_status = (
            "CURRENCY SYMBOL FOUND - CODE NOT IDENTIFIED"
        )

    elif evidence_currencies:

        validation_status = (
            "CURRENCY EVIDENCE FOUND - PRIMARY CURRENCY NOT EXTRACTED"
        )

    elif ocr_evidence:

        validation_status = (
            "CURRENCY NOT EXTRACTED - OCR EVIDENCE AVAILABLE"
        )

    else:

        validation_status = "CURRENCY NOT AVAILABLE"

    # ---------------------------------------------------------
    # Store result
    # ---------------------------------------------------------
    rows.append({
        "case_number": case,
        "primary_currency_values": " | ".join(
            primary_values
        ),
        "primary_source_fields": " | ".join(
            primary_fields
        ),
        "detected_currencies": ", ".join(
            sorted(primary_currencies)
        ),
        "confidence": ", ".join(
            str(c)
            for c in primary_confidences
            if c != ""
        ),
        "currency_evidence_fields": " | ".join(
            evidence_fields
        ),
        "currency_evidence": " | ".join(
            evidence_values
        ),
        "ocr_evidence": ocr_evidence,
        "validation_status": validation_status,
    })


# -------------------------------------------------------------
# Create Excel
# -------------------------------------------------------------
df = pd.DataFrame(rows)

output_file = OUTPUT / "POP_currency_validation.xlsx"

df.to_excel(
    output_file,
    index=False,
    sheet_name="Currency Validation"
)


# -------------------------------------------------------------
# Formatting
# -------------------------------------------------------------
wb = load_workbook(output_file)

ws = wb["Currency Validation"]

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions


for cell in ws[1]:

    cell.font = Font(bold=True)

    cell.alignment = Alignment(
        horizontal="center"
    )


for column in ws.columns:

    max_length = 0

    column_letter = column[0].column_letter

    for cell in column:

        if cell.value is not None:

            max_length = max(
                max_length,
                len(str(cell.value))
            )

    ws.column_dimensions[
        column_letter
    ].width = min(
        max(max_length + 2, 12),
        60
    )


# -------------------------------------------------------------
# Highlight validation status
# -------------------------------------------------------------
status_column = 9

for row in range(
    2,
    ws.max_row + 1
):

    status_cell = ws.cell(
        row=row,
        column=status_column
    )

    status = str(
        status_cell.value
    )

    if status == "VALID CURRENCY":

        status_cell.fill = PatternFill(
            fill_type="solid",
            fgColor="C6EFCE"
        )

    elif (
        "NOT AVAILABLE" in status
        or "NOT EXTRACTED" in status
    ):

        status_cell.fill = PatternFill(
            fill_type="solid",
            fgColor="FFF2CC"
        )

    elif (
        "INVALID" in status
        or "CODE NOT IDENTIFIED" in status
    ):

        status_cell.fill = PatternFill(
            fill_type="solid",
            fgColor="FFC7CE"
        )


wb.save(output_file)


# -------------------------------------------------------------
# Summary
# -------------------------------------------------------------
print("=" * 70)
print("CURRENCY VALIDATION EXCEL CREATED")
print("=" * 70)

print(f"Total cases       : {len(df)}")

print(
    f"Valid currency    : "
    f"{(df['validation_status'] == 'VALID CURRENCY').sum()}"
)

print(
    f"Invalid currency  : "
    f"{(df['validation_status'] == 'INVALID CURRENCY CODE').sum()}"
)

print(
    f"Not detected      : "
    f"{df['validation_status'].str.contains(
        'NOT AVAILABLE|NOT EXTRACTED|PRIMARY CURRENCY NOT EXTRACTED'
    ).sum()}"
)

print(f"Output            : {output_file}")