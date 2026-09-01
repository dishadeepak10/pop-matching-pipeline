import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ============================================================
# PROJECT PATHS
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent

OUTPUT_ROOT = (
    PROJECT_ROOT
    / "data"
    / "output"
)

EXCEL_OUTPUT_PATH = (
    OUTPUT_ROOT
    / "POP_extraction_report.xlsx"
)


# ============================================================
# HELPERS
# ============================================================

def flatten_dict(data, parent_key=""):
    """
    Convert nested JSON objects into a flat dictionary.

    Example:
        {
            "payment_details": {
                "amount": "25000",
                "currency": "AED"
            }
        }

    becomes:

        {
            "payment_details_amount": "25000",
            "payment_details_currency": "AED"
        }
    """

    flattened = {}

    for key, value in data.items():

        new_key = (
            f"{parent_key}_{key}"
            if parent_key
            else key
        )

        if isinstance(value, dict):

            flattened.update(
                flatten_dict(
                    value,
                    new_key
                )
            )

        elif isinstance(value, list):

            flattened[new_key] = "\n".join(
                str(item)
                for item in value
            )

        else:

            flattened[new_key] = value

    return flattened


# ============================================================
# LOAD JSON FILES
# ============================================================

def load_extracted_data():

    json_files = sorted(
        OUTPUT_ROOT.glob(
            "*/extracted.json"
        )
    )

    print(
        f"Found {len(json_files)} extracted JSON files."
    )

    if not json_files:

        raise FileNotFoundError(
            "No extracted.json files found."
        )

    records = []

    for json_file in json_files:

        print(
            f"Reading: {json_file.parent.name}"
        )

        with open(
            json_file,
            "r",
            encoding="utf-8"
        ) as file:

            data = json.load(file)

        flattened_data = flatten_dict(data)

        record = {
            "document_id": json_file.parent.name
        }

        record.update(flattened_data)

        records.append(record)

    return records


# ============================================================
# CREATE EXCEL REPORT
# ============================================================

def create_excel_report(records):

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "POP Extraction"

    # Collect every field found across all documents
    all_columns = []

    for record in records:

        for key in record.keys():

            if key not in all_columns:
                all_columns.append(key)

    # Write headers
    for column_number, column_name in enumerate(
        all_columns,
        start=1
    ):

        cell = worksheet.cell(
            row=1,
            column=column_number,
            value=column_name
        )

        cell.font = cell.font.copy(
            bold=True
        )

    # Write data
    for row_number, record in enumerate(
        records,
        start=2
    ):

        for column_number, column_name in enumerate(
            all_columns,
            start=1
        ):

            value = record.get(
                column_name,
                ""
            )

            worksheet.cell(
                row=row_number,
                column=column_number,
                value=value
            )

    # Freeze header row
    worksheet.freeze_panes = "A2"

    # Auto-size columns
    for column_number in range(
        1,
        worksheet.max_column + 1
    ):

        column_letter = get_column_letter(
            column_number
        )

        max_length = 0

        for cell in worksheet[column_letter]:

            if cell.value is not None:

                length = len(
                    str(cell.value)
                )

                if length > max_length:
                    max_length = length

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            50
        )

    workbook.save(
        EXCEL_OUTPUT_PATH
    )

    print()
    print("=" * 70)
    print("EXCEL REPORT CREATED")
    print("=" * 70)

    print(
        f"Rows: {len(records)}"
    )

    print(
        f"Columns: {len(all_columns)}"
    )

    print(
        f"Output: {EXCEL_OUTPUT_PATH}"
    )


# ============================================================
# MAIN
# ============================================================

def main():

    print()
    print("=" * 70)
    print("POP EXCEL REPORT GENERATION")
    print("=" * 70)

    records = load_extracted_data()

    create_excel_report(records)

    print()
    print("REPORT GENERATION COMPLETE.")


if __name__ == "__main__":
    main()