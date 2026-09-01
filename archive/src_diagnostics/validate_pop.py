import json
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment


ROOT = Path(".")
OUTPUT = ROOT / "data" / "output"
MERGED = OUTPUT / "POP_email_merged_final.xlsx"
REPORT = OUTPUT / "POP_validation_report.xlsx"


# ============================================================
# HELPERS
# ============================================================

def load_fields(case_number):
    path = OUTPUT / f"{case_number}_POP_Document" / "extracted.json"

    if not path.exists():
        return []

    data = json.loads(path.read_text(encoding="utf-8"))

    return data.get("fields", [])


def parse_date(value):
    if not value:
        return None

    text = str(value).strip()

    # Remove common surrounding text
    text = text.replace(" :unselected:", "").strip()

    formats = [
        "%Y/%m/%d",
        "%Y/%m/%d - %I:%M %p",
        "%d-%b-%y",
        "%d-%b-%Y",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%d/%m/%y",
        "%d-%m-%y",
        "%d %b %Y",
        "%d %B %Y",
        "%b %d, %Y",
        "%B %d, %Y",
        "%Y-%m-%d",
        "%d-%m-%Y %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
    ]

    # Handle things such as "24/07/ 2026"
    text = re.sub(r"/\s+", "/", text)

    # Try complete datetime/date formats
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    # Extract date portion from datetime strings
    patterns = [
        r"\b\d{4}/\d{1,2}/\d{1,2}\b",
        r"\b\d{1,2}-[A-Za-z]{3}-\d{2,4}\b",
        r"\b\d{1,2}/\d{1,2}/\d{2,4}\b",
        r"\b\d{1,2}-\d{1,2}-\d{2,4}\b",
        r"\b[A-Za-z]{3,9}\s+\d{1,2},\s+\d{4}\b",
        r"\b\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4}\b",
    ]

    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            candidate = match.group(0)
            for fmt in formats:
                try:
                    return datetime.strptime(candidate, fmt).date()
                except ValueError:
                    pass

    return None


def date_candidates(fields):
    preferred = [
        "transaction_date",
        "transfer_date",
        "wire_date",
        "payment_date",
        "receipt_date",
        "date",
        "effective_date",
        "date_of_application",
        "value_date",
    ]

    result = []

    for field in fields:
        name = str(field.get("field_name", "")).strip()
        value = field.get("value")

        if name in preferred and value not in (None, "", "None"):
            parsed = parse_date(value)

            result.append({
                "field": name,
                "value": str(value),
                "parsed": parsed,
                "confidence": field.get("confidence")
            })

    return result


def normalize_currency(value):
    if value in (None, ""):
        return None

    text = str(value).upper()

    currencies = [
        "AED", "USD", "AUD", "SAR", "GBP",
        "EUR", "USDT", "INR", "PHP"
    ]

    for currency in currencies:
        if re.search(rf"\b{re.escape(currency)}\b", text):
            return currency

    # Dhs / Dirham
    if "DHS" in text or "DIRHAM" in text:
        return "AED"

    # Symbols
    if "$" in text:
        return "USD"

    return None


def clean_amount(value):
    if value in (None, ""):
        return None

    return str(value).strip()


# ============================================================
# LOAD CURRENT EXCEL
# ============================================================

wb_old = load_workbook(MERGED, data_only=True)
ws_old = wb_old.active

headers = [c.value for c in ws_old[1]]
records = []

for row in ws_old.iter_rows(min_row=2, values_only=True):
    records.append(dict(zip(headers, row)))


# ============================================================
# DATE VALIDATION
# ============================================================

date_rows = []

for record in records:

    case = str(record.get("case_number"))

    excel_date = record.get("transaction_date")
    excel_parsed = parse_date(excel_date)

    fields = load_fields(case)
    candidates = date_candidates(fields)

    candidate_text = "; ".join(
        f"{x['field']}={x['value']}"
        for x in candidates
    )

    matching = [
        x for x in candidates
        if excel_parsed is not None
        and x["parsed"] is not None
        and x["parsed"] == excel_parsed
    ]

    notes = ""
    status = ""

    if excel_parsed is None:

        if not candidates:
            status = "UNAVAILABLE"
            notes = "No usable date field found in extracted JSON."
        else:
            status = "REVIEW"
            notes = "Source contains date candidate(s), but Excel has no transaction date."

    else:

        if matching:

            # If multiple dates exist, document which one matched.
            matched_names = ", ".join(x["field"] for x in matching)

            status = "VALID"
            notes = f"Excel date matches source field: {matched_names}"

        elif candidates:

            status = "REVIEW"
            notes = "Excel date does not match the preferred source date candidates."

        else:

            status = "REVIEW"
            notes = "Excel contains a date, but no matching date field was found in extracted JSON."

    date_rows.append([
        case,
        excel_date,
        candidate_text,
        ", ".join(x["field"] for x in matching),
        status,
        notes
    ])


# ============================================================
# CURRENCY VALIDATION
# ============================================================

currency_rows = []

for record in records:

    case = str(record.get("case_number"))

    amount = record.get("amount")
    currency = record.get("currency")

    pop_amount = record.get("pop_amount")
    pop_currency = record.get("pop_currency")

    original_amount = record.get("pop_original_amount")
    original_currency = record.get("pop_original_currency")

    embedded_amount_currency = normalize_currency(amount)
    embedded_pop_currency = normalize_currency(pop_amount)

    notes = ""
    status = "REVIEW"

    # Strong cases where explicit currency agrees with amount.
    if currency and embedded_amount_currency:
        if str(currency).upper() == embedded_amount_currency:
            status = "VALID"
            notes = "Explicit currency agrees with currency embedded in amount."

        else:
            status = "REVIEW"
            notes = (
                f"Currency conflict: column={currency}, "
                f"amount contains={embedded_amount_currency}."
            )

    elif currency and not embedded_amount_currency:
        status = "REVIEW"
        notes = "Currency exists separately; amount contains no explicit currency."

    elif not currency and embedded_amount_currency:
        status = "REVIEW"
        notes = (
            f"Currency is embedded in amount ({embedded_amount_currency}) "
            "but Excel currency column is empty."
        )

    elif pop_currency or original_currency:
        status = "REVIEW"
        notes = "Source contains currency information requiring semantic validation."

    else:
        status = "REVIEW"
        notes = "No reliable currency available in final record."

    currency_rows.append([
        case,
        amount,
        currency,
        pop_amount,
        pop_currency,
        original_amount,
        original_currency,
        status,
        notes
    ])


# ============================================================
# SUMMARY
# ============================================================

date_valid = sum(1 for r in date_rows if r[4] == "VALID")
date_review = sum(1 for r in date_rows if r[4] == "REVIEW")
date_unavailable = sum(1 for r in date_rows if r[4] == "UNAVAILABLE")

currency_valid = sum(1 for r in currency_rows if r[7] == "VALID")
currency_review = sum(1 for r in currency_rows if r[7] == "REVIEW")

# Read confidence from extracted JSON
confidence_values = []

for record in records:
    case = str(record.get("case_number"))
    fields = load_fields(case)

    for field in fields:
        confidence = field.get("confidence")

        if isinstance(confidence, (int, float)):
            confidence_values.append(float(confidence))

average_field_confidence = (
    sum(confidence_values) / len(confidence_values)
    if confidence_values
    else None
)


# ============================================================
# CREATE REPORT
# ============================================================

wb = Workbook()

# ------------------------------------------------------------
# SUMMARY
# ------------------------------------------------------------

ws = wb.active
ws.title = "Summary"

summary = [
    ["POP VALIDATION SUMMARY", ""],
    ["Total cases", len(records)],
    ["Date validation - VALID", date_valid],
    ["Date validation - REVIEW", date_review],
    ["Date validation - UNAVAILABLE", date_unavailable],
    ["Currency validation - VALID", currency_valid],
    ["Currency validation - REVIEW", currency_review],
    ["Average extracted field confidence", round(average_field_confidence, 2) if average_field_confidence else None],
    ["Processing failures in final merge", 0],
    ["Current stage", "Validation before case matching"],
]

for row in summary:
    ws.append(row)


# ------------------------------------------------------------
# DATE VALIDATION
# ------------------------------------------------------------

ws_date = wb.create_sheet("Date Validation")

ws_date.append([
    "Case Number",
    "Excel Transaction Date",
    "Source Date Candidates",
    "Matched Source Field",
    "Status",
    "Notes"
])

for row in date_rows:
    ws_date.append(row)


# ------------------------------------------------------------
# CURRENCY VALIDATION
# ------------------------------------------------------------

ws_cur = wb.create_sheet("Currency Validation")

ws_cur.append([
    "Case Number",
    "Excel Amount",
    "Excel Currency",
    "POP Amount",
    "POP Currency",
    "Original Amount",
    "Original Currency",
    "Status",
    "Notes"
])

for row in currency_rows:
    ws_cur.append(row)


# ============================================================
# FORMATTING
# ============================================================

for sheet in wb.worksheets:

    sheet.freeze_panes = "A2"

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for column in sheet.columns:

        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        sheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            60
        )

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )


wb.save(REPORT)

print()
print("=" * 70)
print("POP VALIDATION REPORT")
print("=" * 70)
print(f"Total cases                 : {len(records)}")
print(f"Date VALID                  : {date_valid}")
print(f"Date REVIEW                 : {date_review}")
print(f"Date UNAVAILABLE            : {date_unavailable}")
print(f"Currency VALID              : {currency_valid}")
print(f"Currency REVIEW             : {currency_review}")
print(
    f"Average field confidence    : "
    f"{average_field_confidence:.2f}"
    if average_field_confidence
    else "Average field confidence    : N/A"
)
print()
print(f"Report created              : {REPORT}")
print("=" * 70)
