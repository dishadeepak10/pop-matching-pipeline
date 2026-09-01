import json
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# PROJECT PATHS
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_ROOT = PROJECT_ROOT / "data" / "output"


# ============================================================
# CASES
# ============================================================

CASES = [
    "00084283", "00084285", "00084308", "00084323", "00084360",
    "00084362", "00084373", "00084375", "00084379", "00084384",
    "00084401", "00084501", "00084572", "00084596", "00084670",
    "00084696", "00084725", "00084741", "00084742", "00084772",
    "00084822", "00084826", "00084851", "00084879", "00084922"
]


# ============================================================
# AMOUNT FIELD DETECTION
# ============================================================

AMOUNT_FIELD_PATTERN = re.compile(
    r"(amount|value|total|fee|charge|price|payment|"
    r"balance|debit|credit|transfer|transaction|"
    r"receipt|wire|tax|cost|commission|"
    r"exchange_rate|conversion_rate)",
    re.IGNORECASE
)


# ============================================================
# CURRENCY / AMOUNT PATTERNS
# ============================================================

CURRENCY_CODES = (
    "AED|USD|EUR|GBP|SAR|INR|AUD|CAD|CHF|"
    "JPY|CNY|HKD|SGD|NZD|USDT"
)

AMOUNT_PATTERN = re.compile(
    rf"""
    (?:
        (?:{CURRENCY_CODES})\s*
        [-+]?
        (?:\d{{1,3}}(?:,\d{{3}})+(?:\.\d+)?|\d+(?:\.\d+)?)
    |
        [-+]?
        (?:\d{{1,3}}(?:,\d{{3}})+(?:\.\d+)?|\d+(?:\.\d+)?)
        \s*
        (?:{CURRENCY_CODES})
    |
        [$€£]\s*
        [-+]?
        (?:\d{{1,3}}(?:,\d{{3}})+(?:\.\d+)?|\d+(?:\.\d+)?)
    |
        [-+]?
        (?:\d{{1,3}}(?:,\d{{3}})+(?:\.\d+)?|\d+(?:\.\d+)?)
    )
    """,
    re.IGNORECASE | re.VERBOSE
)


# ============================================================
# LOAD JSON
# ============================================================

def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ============================================================
# EXTRACT AMOUNT EVIDENCE FROM VALUE
# ============================================================

def extract_amount_evidence(value):
    if value is None:
        return []

    text = str(value)

    matches = AMOUNT_PATTERN.findall(text)

    return [
        m.strip()
        for m in matches
        if m.strip()
    ]


# ============================================================
# VALIDATE ONE DOCUMENT
# ============================================================

def validate_document(case_number):

    folder = OUTPUT_ROOT / f"{case_number}_POP_Document"

    extracted_path = folder / "extracted.json"
    ocr_path = folder / "ocr.txt"

    primary_fields = []
    evidence_fields = []
    detected_amounts = []

    ocr_evidence = []

    # --------------------------------------------------------
    # EXTRACTED JSON
    # --------------------------------------------------------

    if extracted_path.exists():

        try:
            data = load_json(extracted_path)
        except Exception as e:

            return {
                "case_number": case_number,
                "detected_amounts": "",
                "primary_source_fields": "",
                "amount_evidence_fields": "",
                "ocr_evidence": "",
                "validation_status": f"INVALID JSON: {e}",
            }

        fields = data.get("fields", [])

        if isinstance(fields, list):

            for field in fields:

                if not isinstance(field, dict):
                    continue

                field_name = str(
                    field.get("field_name", "")
                ).strip()

                value = field.get("value")

                if not field_name:
                    continue

                # ------------------------------------------------
                # Only inspect fields likely related to amounts
                # ------------------------------------------------

                if not AMOUNT_FIELD_PATTERN.search(field_name):
                    continue

                if value in (
                    None,
                    "",
                    "N/A",
                    "Not Applicable"
                ):
                    continue

                evidence = extract_amount_evidence(value)

                if evidence:

                    for amount in evidence:

                        if amount not in detected_amounts:
                            detected_amounts.append(amount)

                    # Fields that actually contain numeric amounts
                    primary_fields.append(field_name)

                elif any(
                    code.lower() in str(value).lower()
                    for code in [
                        "AED", "USD", "EUR", "GBP",
                        "SAR", "INR", "AUD", "USDT",
                        "$", "€", "£"
                    ]
                ):

                    evidence_fields.append(field_name)

        # --------------------------------------------------------
        # Remove duplicates while preserving order
        # --------------------------------------------------------

        primary_fields = list(
            dict.fromkeys(primary_fields)
        )

        evidence_fields = list(
            dict.fromkeys(evidence_fields)
        )

    # ============================================================
    # OCR EVIDENCE
    # ============================================================

    if ocr_path.exists():

        try:
            ocr_lines = ocr_path.read_text(
                encoding="utf-8",
                errors="replace"
            ).splitlines()

            for line in ocr_lines:

                line = line.strip()

                if not line:
                    continue

                if AMOUNT_PATTERN.search(line):

                    ocr_evidence.append(line)

        except Exception:
            pass

    # Keep OCR evidence concise
    ocr_evidence = ocr_evidence[:8]

    # ============================================================
    # VALIDATION STATUS
    # ============================================================

    if detected_amounts:

        status = "VALID AMOUNT"

    elif primary_fields or evidence_fields:

        status = "AMOUNT EVIDENCE FOUND - REVIEW"

    elif ocr_evidence:

        status = "AMOUNT FOUND IN OCR - NOT EXTRACTED"

    else:

        status = "AMOUNT NOT AVAILABLE"

    return {
        "case_number": case_number,
        "detected_amounts": ", ".join(detected_amounts),
        "primary_source_fields": " | ".join(primary_fields),
        "amount_evidence_fields": " | ".join(evidence_fields),
        "ocr_evidence": " | ".join(ocr_evidence),
        "validation_status": status,
    }


# ============================================================
# MAIN
# ============================================================

def main():

    print("=" * 70)
    print("AMOUNT VALIDATION")
    print("=" * 70)

    rows = []

    for case_number in CASES:

        result = validate_document(case_number)

        rows.append(result)

        print()
        print(
            f"{case_number}: "
            f"{result['validation_status']}"
        )

        if result["detected_amounts"]:

            print(
                f"  amounts: "
                f"{result['detected_amounts']}"
            )

        if result["primary_source_fields"]:

            print(
                f"  fields: "
                f"{result['primary_source_fields']}"
            )

    # ========================================================
    # DATAFRAME
    # ========================================================

    df = pd.DataFrame(rows)

    output_file = (
        OUTPUT_ROOT /
        "POP_amount_validation.xlsx"
    )

    df.to_excel(
        output_file,
        index=False,
        sheet_name="Amount Validation"
    )

    # ========================================================
    # FORMAT EXCEL
    # ========================================================

    wb = load_workbook(output_file)

    ws = wb["Amount Validation"]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Header
    for cell in ws[1]:

        cell.font = Font(bold=True)

        cell.alignment = Alignment(
            horizontal="center"
        )

    # Column widths
    for column in ws.columns:

        max_length = 0

        column_letter = (
            column[0].column_letter
        )

        for cell in column:

            if cell.value is not None:

                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[
            column_letter
        ].width = min(
            max(max_length + 2, 12),
            80
        )

    # ========================================================
    # STATUS HIGHLIGHTING
    # ========================================================

    status_column = None

    for cell in ws[1]:

        if cell.value == "validation_status":

            status_column = cell.column

    if status_column:

        for row in range(
            2,
            ws.max_row + 1
        ):

            cell = ws.cell(
                row=row,
                column=status_column
            )

            status = str(cell.value)

            if status == "VALID AMOUNT":

                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="C6EFCE"
                )

            elif "REVIEW" in status:

                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="FFF2CC"
                )

            elif "NOT AVAILABLE" in status:

                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="F4CCCC"
                )

    wb.save(output_file)

    # ========================================================
    # SUMMARY
    # ========================================================

    valid_count = (
        df["validation_status"]
        .eq("VALID AMOUNT")
        .sum()
    )

    review_count = (
        df["validation_status"]
        .str.contains(
            "REVIEW",
            na=False
        )
        .sum()
    )

    ocr_only_count = (
        df["validation_status"]
        .eq(
            "AMOUNT FOUND IN OCR - NOT EXTRACTED"
        )
        .sum()
    )

    unavailable_count = (
        df["validation_status"]
        .eq("AMOUNT NOT AVAILABLE")
        .sum()
    )

    print()
    print("=" * 70)
    print("AMOUNT VALIDATION EXCEL CREATED")
    print("=" * 70)

    print(
        f"Total cases       : {len(df)}"
    )

    print(
        f"Valid amount      : {valid_count}"
    )

    print(
        f"Needs review      : {review_count}"
    )

    print(
        f"OCR only          : {ocr_only_count}"
    )

    print(
        f"Not detected      : {unavailable_count}"
    )

    print(
        f"Output            : {output_file}"
    )

    print("=" * 70)


if __name__ == "__main__":
    main()