import re
from pathlib import Path
from openpyxl import load_workbook


# ============================================================
# PATHS
# ============================================================

EMAIL_LOG = Path(
    r".\data\input\Disha_Learning\Disha_Learning\_POP_EmailsLog.xlsx"
)

OUTPUT_DIR = Path(r".\data\output")


# ============================================================
# PILOT CASES
# ============================================================

PILOT_CASES = [
    "00084922",
    "00084879",
    "00084826",
]


# ============================================================
# LOAD EMAIL DATA
# ============================================================

def load_email_records():
    workbook = load_workbook(EMAIL_LOG, data_only=True)
    sheet = workbook["POP_attachments"]

    headers = [cell.value for cell in sheet[1]]

    records = {}

    for row in sheet.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        record = dict(zip(headers, values))

        case_number = str(record.get("Case Number") or "").strip()

        if case_number:
            records[case_number] = record

    return records


# ============================================================
# EMAIL FIELD EXTRACTION
# ============================================================

def extract_email_fields(email_body):
    text = str(email_body or "")

    def extract(pattern):
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        return match.group(1).strip() if match else None

    return {
        "email_case_number": extract(
            r"Case Number:\s*([0-9]+)"
        ),

        "email_created_date": extract(
            r"Created Date:\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})"
        ),
        "email_receipt_reference": extract(
            r"Receipt Acknowledgement:\s*([A-Za-z0-9\-]+?)(?=Receipt Amount:)"
        ),
        "email_receipt_amount": extract(
            r"Receipt Amount:\s*([0-9,]+(?:\.[0-9]+)?)"
        ),
        "email_bank_name": extract(
            r"Bank Name:\s*(.*?)(?=Payment Method:)",
        ),
        "email_payment_method": extract(
            r"Payment Method:\s*(.*?)(?=Customer Last Name:)"
        ),
        "email_customer_name": extract(
            r"Customer Last Name:\s*(.*?)(?=Bank\s*Account\s*Number:)"
        ),
        "email_bank_account": extract(
            r"Bank\s*Account\s*Number:\s*(.*?)(?=Remarks:)"
        ),
        "email_customer_bank": extract(
            r"Remarks:\s*Customer Bank Name:\s*(.*?)(?=\n|CREATOR INFORMATION)"
        ),
    }


# ============================================================
# LOAD POP JSON
# ============================================================

def load_pop(case_number):
    json_path = (
        OUTPUT_DIR
        / f"{case_number}_POP_Document"
        / "extracted.json"
    )

    import json

    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


# ============================================================
# PILOT MERGE
# ============================================================

def build_pilot_record(case_number, email_record, pop_data):

    email_fields = extract_email_fields(
        email_record.get("EmailBody")
    )

    # --------------------------------------------------------
    # Helper
    # --------------------------------------------------------

    def first_value(*keys):
        for key in keys:
            value = pop_data.get(key)
            if value not in (None, ""):
                return value
        return None

    # --------------------------------------------------------
    # POP important fields
    # --------------------------------------------------------

    sender_name = first_value(
        "sender_name",
        "account_holder_name",
        "payer_name",
        "applicant_name",
    )

    transaction_date = first_value(
        "transaction_date",
        "payment_date",
        "receipt_date",
        "date_of_transaction",
    )

    reference_number = first_value(
        "reference_number",
        "transaction_reference",
        "transaction_number",
        "booking_reference",
    )

    sender_bank = first_value(
        "sender_bank",
        "payer_bank",
        "remitter_bank",
        "bank_name",
    )

    beneficiary_name = first_value(
        "beneficiary_name",
        "recipient_name",
        "counterparty_name",
    )

    beneficiary_bank = first_value(
        "beneficiary_bank_name",
        "beneficiary_bank",
        "recipient_bank_name",
    )

    # --------------------------------------------------------
    # Amount logic
    #
    # Prefer original transaction amount when available.
    # Otherwise use transfer/POP amount.
    # --------------------------------------------------------

    pop_original_amount = pop_data.get("original_amount")
    pop_original_currency = pop_data.get("original_currency")

    pop_amount = first_value(
        "transfer_amount",
        "amount",
        "amount_sent",
        "total_amount",
        "total_amount_debited",
    )

    pop_currency = first_value(
        "transfer_currency",
        "currency",
        "currency_code",
    )

    # If original amount + original currency exist,
    # treat those as the primary transaction amount.
    if pop_original_amount not in (None, ""):
        amount = pop_original_amount
        currency = pop_original_currency or pop_currency
        amount_source = "POP original amount"
    else:
        amount = pop_amount
        currency = pop_currency

        if amount not in (None, ""):
            amount_source = "POP"
        else:
            amount_source = None
        # --------------------------------------------------------
    # EMAIL FALLBACKS FOR IMPORTANT FIELDS
    # --------------------------------------------------------

    if not sender_name:
        sender_name = email_fields.get("email_customer_name")
        if sender_name:
            sender_name_source = "EMAIL"
        else:
            sender_name_source = None
    else:
        sender_name_source = "POP"

    if not reference_number:
        reference_number = email_fields.get(
            "email_receipt_reference"
        )
        if reference_number:
            reference_source = "EMAIL"
        else:
            reference_source = None
    else:
        reference_source = "POP"
    # --------------------------------------------------------
    # Final prioritized record
    # --------------------------------------------------------
    
    return {
        "case_number": case_number,

        # ====================================================
        # PRIORITIZED FIELDS
        # ====================================================

        "sender_name": sender_name,
        "sender_name_source": sender_name_source,

        "transaction_date": transaction_date,

        "amount": amount,

        "currency": currency,

        "reference_number": reference_number,

        "sender_bank": sender_bank,

        "beneficiary_name": beneficiary_name,

        "beneficiary_bank": beneficiary_bank,

        # ====================================================
        # SOURCE / TRACEABILITY
        # ====================================================

        "amount_source": amount_source,

        "date_source": (
            "POP"
            if transaction_date
            else "Not available in POP/OCR"
        ),

        "reference_source": reference_source,
        # ====================================================
        # POP SUPPORTING FIELDS
        # ====================================================

        "pop_amount": pop_amount,

        "pop_currency": pop_currency,

        "pop_original_amount": pop_original_amount,

        "pop_original_currency": pop_original_currency,

        "pop_exchange_rate": pop_data.get(
            "exchange_rate"
        ),

        "pop_value_date": pop_data.get(
            "value_date"
        ),

        "pop_booking_reference": pop_data.get(
            "booking_reference"
        ),

        # ====================================================
        # EMAIL FIELDS
        # ====================================================

        **email_fields,
    }
# ============================================================
# MAIN
# ============================================================

def main():

    print("=" * 100)
    print("EMAIL + POP PILOT MERGE")
    print("=" * 100)

    email_records = load_email_records()

    for case_number in PILOT_CASES:

        print("\n" + "=" * 100)
        print(f"CASE: {case_number}")
        print("=" * 100)

        email_record = email_records.get(case_number)

        if not email_record:
            print("ERROR: Email record not found")
            continue

        pop_data = load_pop(case_number)

        result = build_pilot_record(
            case_number,
            email_record,
            pop_data,
        )

        for key, value in result.items():
            print(f"{key:30} : {value}")


if __name__ == "__main__":
    main()