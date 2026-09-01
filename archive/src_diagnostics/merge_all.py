from services.pop_service import load_pop
from generate_report import normalize_record
from services.email_service import (
    load_email_records,
    extract_email_fields,
)

import json
import re
import logging
from pathlib import Path

from utils.ocr_utils import load_ocr, extract_ocr_date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# LOGGING
# ============================================================

LOG_FILE = Path(
    r".\data\output\merge_all.log"
)

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.ERROR,
    format=(
        "%(asctime)s | "
        "%(levelname)s | "
        "%(message)s"
    ),
)

logger = logging.getLogger(__name__)


# ============================================================
# AMOUNT / CURRENCY NORMALIZATION
# ============================================================

def normalize_amount_currency(amount, currency=None):

    if amount in (None, ""):
        return amount, currency

    text = str(amount).strip()
    detected_currency = None

    # --------------------------------------------------------
    # 1. Detect currency codes inside the amount
    # --------------------------------------------------------

    currency_pattern = (
        r"\b(?:AED|AUD|SAR|GBP|EUR|USD|USDT)\b"
    )

    matches = re.findall(
        currency_pattern,
        text,
        re.IGNORECASE
    )

    if matches:

        detected_currency = matches[0].upper()

        # Remove currency code from amount
        text = re.sub(
            currency_pattern,
            "",
            text,
            flags=re.IGNORECASE
        )

    # --------------------------------------------------------
    # 2. Handle Dhs / Dirham notation
    # --------------------------------------------------------

    if re.search(
        r"\bDhs\b|Dhs",
        text,
        re.IGNORECASE
    ):

        detected_currency = "AED"

        text = re.sub(
            r"\bDhs\b|Dhs",
            "",
            text,
            flags=re.IGNORECASE
        )

    # --------------------------------------------------------
    # 3. Remove descriptive text in parentheses
    # --------------------------------------------------------

    text = re.sub(
        r"\(.*?\)",
        "",
        text
    )

    text = text.strip()

    # --------------------------------------------------------
    # 4. Detect trailing minus sign
    #
    # Example:
    #     51000-  -> -51000
    # --------------------------------------------------------

    trailing_negative = bool(
        re.search(
            r"-\s*$",
            text
        )
    )

    text = re.sub(
        r"-\s*$",
        "",
        text
    )

    # --------------------------------------------------------
    # 5. Remove unwanted characters
    #
    # Keep:
    # digits
    # decimal separators
    # minus sign
    # --------------------------------------------------------

    text = re.sub(
        r"[^\d,.\-]",
        "",
        text
    )

    text = text.strip()

    if not text:

        return amount, (
            detected_currency
            if detected_currency
            else currency
        )

    # --------------------------------------------------------
    # 6. Convert number formatting
    #
    # Supported:
    #
    # 100,000.00  -> 100000.00
    # 50.000,00   -> 50000.00
    # 12.101,03   -> 12101.03
    # 50000       -> 50000.00
    # --------------------------------------------------------

    if "," in text and "." in text:

        last_comma = text.rfind(",")
        last_dot = text.rfind(".")

        # European format:
        # 50.000,00

        if last_comma > last_dot:

            text = text.replace(
                ".",
                ""
            )

            text = text.replace(
                ",",
                "."
            )

        # Standard format:
        # 50,000.00

        else:

            text = text.replace(
                ",",
                ""
            )

    elif "," in text:

        parts = text.split(",")

        # Example:
        # 50,000 -> thousands separator

        if (
            len(parts) == 2
            and len(parts[1]) == 3
            and parts[0].replace("-", "").isdigit()
        ):

            text = "".join(parts)

        # Example:
        # 50,000,000

        elif all(
            len(part) == 3
            for part in parts[1:]
        ):

            text = "".join(parts)

        # Example:
        # 50,00 -> decimal comma

        else:

            text = text.replace(
                ",",
                "."
            )

    elif "." in text:

        parts = text.split(".")

        # Example:
        # 50.000 -> likely thousands separator

        if (
            len(parts) == 2
            and len(parts[1]) == 3
            and parts[0].replace("-", "").isdigit()
        ):

            text = "".join(parts)

    # --------------------------------------------------------
    # 7. Apply trailing negative sign
    # --------------------------------------------------------

    if (
        trailing_negative
        and not text.startswith("-")
    ):

        text = "-" + text

    # --------------------------------------------------------
    # 8. Convert to a clean numeric string
    # --------------------------------------------------------

    try:

        numeric_value = float(text)

        text = f"{numeric_value:.2f}"

    except ValueError:

        # If something unexpected survives,
        # return the cleaned text instead of crashing.

        pass

    # --------------------------------------------------------
    # 9. Currency priority
    #
    # Currency explicitly embedded in the amount
    # is stronger than the existing currency field.
    # --------------------------------------------------------

    if detected_currency:

        currency = detected_currency

    elif isinstance(currency, str):

        currency_match = re.fullmatch(
            r"\s*(AED|AUD|SAR|GBP|EUR|USD|USDT)\s*",
            currency,
            re.IGNORECASE
        )

        if currency_match:

            currency = (
                currency_match.group(1).upper()
            )

        else:

            # If the currency field contains a valid
            # currency somewhere inside malformed OCR text,
            # recover it.

            fallback_currency = re.search(
                r"\b(AED|AUD|SAR|GBP|EUR|USD|USDT)\b",
                currency,
                re.IGNORECASE
            )

            if fallback_currency:

                currency = (
                    fallback_currency
                    .group(1)
                    .upper()
                )

            else:

                currency = None

    return text, currency


# ============================================================
# PATHS
# ============================================================

EMAIL_LOG = Path(
    r".\data\input\Disha_Learning\Disha_Learning\_POP_EmailsLog.xlsx"
)

OUTPUT_DIR = Path(
    r".\data\output"
)

OUTPUT_EXCEL = (
    OUTPUT_DIR
    / "POP_email_merged_final.xlsx"
)


# ============================================================
# MERGE ONE CASE
# ============================================================

def build_record(
    case_number,
    email_record,
    pop_data,
    ocr_text
):

    email_fields = extract_email_fields(
        email_record.get("EmailBody")
    )

    def first_value(*keys):

        for key in keys:

            value = pop_data.get(key)

            if value not in (None, ""):

                return value

        return None


    # ========================================================
    # SENDER
    # ========================================================

    sender_name = first_value(
        "sender_name",
        "account_holder_name",
        "payer_name",
        "applicant_name",
    )

    if sender_name:

        sender_name_source = "POP"

    else:

        sender_name = email_fields.get(
            "email_customer_name"
        )

        sender_name_source = (
            "EMAIL"
            if sender_name
            else None
        )


    # ========================================================
    # DATE
    # ========================================================

    transaction_date = first_value(
        "transaction_date",
        "payment_date",
        "receipt_date",
        "date_of_transaction",
    )

    if transaction_date:

        date_source = "POP"

    else:

        transaction_date = extract_ocr_date(
            ocr_text
        )

        date_source = (
            "RAW OCR"
            if transaction_date
            else "Not available in POP/OCR"
        )


    # ========================================================
    # REFERENCE
    # ========================================================

    reference_number = first_value(
        "reference_number",
        "transaction_reference",
        "transaction_number",
        "booking_reference",
    )

    if reference_number:

        reference_source = "POP"

    else:

        reference_number = email_fields.get(
            "email_receipt_reference"
        )

        reference_source = (
            "EMAIL"
            if reference_number
            else None
        )


    # ========================================================
    # BANK
    # ========================================================

    sender_bank = first_value(
        "sender_bank",
        "payer_bank",
        "remitter_bank",
        "bank_name",
    )


    # ========================================================
    # BENEFICIARY
    # ========================================================

    beneficiary_name = first_value(
        "beneficiary_name",
        "recipient_name",
        "counterparty_name",
    )

    beneficiary_bank = first_value(
        "beneficiary_bank_name",
        "beneficiary_bank",
        "recipient_bank_name",
    )


    # ========================================================
    # AMOUNT
    # ========================================================

    pop_original_amount = pop_data.get(
        "original_amount"
    )

    pop_original_currency = pop_data.get(
        "original_currency"
    )

    pop_amount = first_value(
        "transfer_amount",
        "amount",
        "amount_sent",
        "total_amount",
        "total_amount_debited",
    )

    pop_currency = first_value(
        "transfer_currency",
        "currency",
        "currency_code",
    )


    # --------------------------------------------------------
    # Normalize POP amount/currency
    # --------------------------------------------------------

    normalized_pop_amount, normalized_pop_currency = (
        normalize_amount_currency(
            pop_amount,
            pop_currency
        )
    )

    # Keep normalized values in supporting POP fields

    pop_amount = normalized_pop_amount
    pop_currency = normalized_pop_currency


    # --------------------------------------------------------
    # Select prioritized amount source
    # --------------------------------------------------------

    if pop_original_amount not in (None, ""):

        amount = pop_original_amount

        currency = (
            pop_original_currency
            or pop_currency
        )

        amount_source = (
            "POP original amount"
        )

    elif pop_amount not in (None, ""):

        amount = pop_amount

        currency = pop_currency

        amount_source = "POP"
    else:

        amount = email_fields.get(
            "email_receipt_amount"
        )

        # Keep POP currency if it exists,
        # even when POP amount itself is missing.
        currency = pop_currency

        amount_source = (
            "EMAIL"
            if amount
            else None
        )


    # --------------------------------------------------------
    # Normalize final prioritized amount/currency
    # --------------------------------------------------------

    amount, currency = normalize_amount_currency(
        amount,
        currency
    )


    # ========================================================
    # RECORD
    # ========================================================

    return {

        # ----------------------------------------------------
        # PRIORITIZED FIELDS
        # ----------------------------------------------------

        "case_number": case_number,

        "sender_name": sender_name,

        "transaction_date": transaction_date,

        "amount": amount,

        "currency": currency,

        "reference_number": reference_number,

        "sender_bank": sender_bank,

        "beneficiary_name": beneficiary_name,

        "beneficiary_bank": beneficiary_bank,


        # ----------------------------------------------------
        # SOURCE TRACKING
        # ----------------------------------------------------

        "sender_name_source": sender_name_source,

        "amount_source": amount_source,

        "date_source": date_source,

        "reference_source": reference_source,


        # ----------------------------------------------------
        # EMAIL SUPPORTING FIELDS
        # ----------------------------------------------------

        **email_fields,


        # ----------------------------------------------------
        # POP SUPPORTING FIELDS
        # ----------------------------------------------------

        "pop_amount": pop_amount,

        "pop_currency": pop_currency,

        "pop_original_amount": (
            pop_original_amount
        ),

        "pop_original_currency": (
            pop_original_currency
        ),

        "pop_exchange_rate": pop_data.get(
            "exchange_rate"
        ),

        "pop_value_date": pop_data.get(
            "value_date"
        ),

        "pop_booking_reference": pop_data.get(
            "booking_reference"
        ),

    }


# ============================================================
# EXCEL SAFE VALUE
# ============================================================

def excel_safe_value(value):

    if value is None:

        return ""

    if isinstance(value, dict):

        return " | ".join(
            f"{key}: {val}"
            for key, val in value.items()
        )

    if isinstance(value, list):

        return " | ".join(
            str(item)
            for item in value
        )

    return value


# ============================================================
# WRITE EXCEL
# ============================================================

def write_excel(records):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Merged POP Data"

    if not records:

        workbook.save(
            OUTPUT_EXCEL
        )

        return


    headers = list(
        records[0].keys()
    )

    sheet.append(headers)


    # --------------------------------------------------------
    # Header formatting
    # --------------------------------------------------------

    for cell in sheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="D9EAF7"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


    # --------------------------------------------------------
    # Data
    # --------------------------------------------------------

    for record in records:

        sheet.append(
            [
                excel_safe_value(
                    record.get(header)
                )
                for header in headers
            ]
        )


    # --------------------------------------------------------
    # Freeze header
    # --------------------------------------------------------

    sheet.freeze_panes = "A2"


    # --------------------------------------------------------
    # Auto width
    # --------------------------------------------------------

    for column in sheet.columns:

        max_length = 0

        column_letter = (
            column[0].column_letter
        )

        for cell in column:

            value = str(
                cell.value or ""
            )

            max_length = max(
                max_length,
                len(value)
            )

        sheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            45
        )


    # --------------------------------------------------------
    # Save
    # --------------------------------------------------------

    workbook.save(
        OUTPUT_EXCEL
    )


# ============================================================
# MAIN
# ============================================================

def main():

    print("=" * 100)
    print("FULL EMAIL + POP MERGE")
    print("=" * 100)


    # --------------------------------------------------------
    # Load email records
    # --------------------------------------------------------

    email_records = load_email_records(
        EMAIL_LOG
    )


    # --------------------------------------------------------
    # Discover all POP JSON files
    # --------------------------------------------------------

    json_files = sorted(
        OUTPUT_DIR.glob(
            "*_POP_Document/extracted.json"
        )
    )


    print(
        f"\nFound {len(json_files)} POP JSON files."
    )


    records = []

    missing_email = []

    failed_cases = []


    # ========================================================
    # PROCESS EACH CASE
    # ========================================================

    for json_path in json_files:

        case_number = (
            json_path.parent.name
            .replace(
                "_POP_Document",
                ""
            )
        )


        print(
            f"Processing: {case_number}"
        )


        try:

            # ------------------------------------------------
            # Load POP JSON
            # ------------------------------------------------

            pop_data = load_pop(
                case_number,
                OUTPUT_DIR
            )


            if pop_data is None:

                print(
                    "  WARNING: POP JSON missing"
                )

                logger.error(
                    "POP JSON missing for case %s",
                    case_number
                )

                failed_cases.append(
                    case_number
                )

                continue


            # ------------------------------------------------
            # Normalize POP data
            # ------------------------------------------------

            pop_data = normalize_record(
                pop_data,
                f"{case_number}_POP_Document"
            )


            # ------------------------------------------------
            # Get email record
            # ------------------------------------------------

            email_record = email_records.get(
                case_number
            )


            if not email_record:

                missing_email.append(
                    case_number
                )

                email_record = {
                    "EmailBody": ""
                }


            # ------------------------------------------------
            # Load OCR
            # ------------------------------------------------

            ocr_text = load_ocr(
                case_number,
                OUTPUT_DIR
            )


            # ------------------------------------------------
            # Build merged record
            # ------------------------------------------------

            record = build_record(
                case_number,
                email_record,
                pop_data,
                ocr_text
            )


            records.append(
                record
            )


            print(
                f"  SUCCESS: {case_number}"
            )


        except Exception as e:

            # ------------------------------------------------
            # IMPORTANT:
            # One failed case does NOT stop the batch.
            # ------------------------------------------------

            failed_cases.append(
                case_number
            )


            logger.exception(
                "Failed processing case %s: %s",
                case_number,
                e
            )


            print(
                f"  ERROR: Failed processing "
                f"{case_number}: {e}"
            )


            # Continue with next case

            continue


    # ========================================================
    # WRITE EXCEL
    # ========================================================

    write_excel(
        records
    )


    # ========================================================
    # FINAL SUMMARY
    # ========================================================

    print(
        "\n" + "=" * 100
    )

    print(
        "MERGE COMPLETE"
    )

    print(
        "=" * 100
    )


    print(
        f"Total POP files : {len(json_files)}"
    )

    print(
        f"Records written : {len(records)}"
    )

    print(
        f"Failed cases    : {len(failed_cases)}"
    )

    print(
        f"Output          : {OUTPUT_EXCEL}"
    )

    print(
        f"Error log       : {LOG_FILE}"
    )


    # --------------------------------------------------------
    # Missing email summary
    # --------------------------------------------------------

    if missing_email:

        print(
            "\nCases without email records:"
        )

        for case in missing_email:

            print(
                f"  {case}"
            )


    # --------------------------------------------------------
    # Failed case summary
    # --------------------------------------------------------

    if failed_cases:

        print(
            "\nCases that failed:"
        )

        for case in failed_cases:

            print(
                f"  {case}"
            )

    else:

        print(
            "\nNo processing failures."
        )


# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == "__main__":

    main()