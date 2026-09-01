import re
from openpyxl import load_workbook


def load_email_records(email_log):
    workbook = load_workbook(
        email_log,
        data_only=True
    )

    sheet = workbook["POP_attachments"]

    headers = [
        cell.value
        for cell in sheet[1]
    ]

    records = {}

    for row in sheet.iter_rows(min_row=2):

        values = [
            cell.value
            for cell in row
        ]

        record = dict(
            zip(headers, values)
        )

        case_number = str(
            record.get("Case Number") or ""
        ).strip()

        if case_number:
            records[case_number] = record

    return records


def extract_email_fields(email_body):

    text = str(email_body or "")

    def extract(pattern):

        match = re.search(
            pattern,
            text,
            re.IGNORECASE
        )

        if match:
            return match.group(1).strip()

        return None

    return {

        "email_case_number": extract(
            r"Case Number:\s*([0-9]+)"
        ),

        "email_created_date": extract(
            r"Created Date:\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})"
        ),

        "email_receipt_reference": extract(
            r"Receipt Acknowledgement:\s*([A-Za-z0-9\-]+)"
        ),

        "email_receipt_amount": extract(
            r"Receipt Amount:\s*([0-9,]+(?:\.[0-9]+)?)"
        ),

        "email_bank_name": extract(
            r"Bank\s*Name:\s*(.*?)(?=Payment\s*Method:)"
        ),

        "email_payment_method": extract(
            r"Payment\s*Method:\s*(.*?)(?=Customer\s*Last\s*Name:)"
        ),

        "email_customer_name": extract(
            r"Customer\s*Last\s*Name:\s*(.*?)(?=Bank\s*Account\s*Number:)"
        ),

        "email_bank_account": extract(
            r"Bank\s*Account\s*Number:\s*(.*?)(?=Remarks:)"
        ),

        "email_customer_bank": extract(
            r"Remarks:\s*Customer\s*Bank\s*Name:\s*(.*?)(?=\n|CREATOR INFORMATION)"
        ),
    }
